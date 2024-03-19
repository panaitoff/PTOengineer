import sys
import sqlite3

from PyQt6.QtCore import QSize
from PyQt6.QtGui import QPixmap, QPainter
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QPushButton, QTableWidget, QTableWidgetItem, QFileDialog, QDialog, QLabel, QLineEdit,
                             QMessageBox)
from PyQt6.QtPrintSupport import QPrintDialog, QPrinter
import openpyxl
import os
import win32com.client
import docx


DOC_path = 'temp_files/done.docx'


class CardForm(QDialog):
    def __init__(self, project_id, project_name):
        super().__init__()

        self.setWindowTitle("Форма для ввода данных")
        self.setGeometry(100, 100, 600, 400)

        connection = sqlite3.connect("projects.db")
        cursor = connection.cursor()
        data = cursor.execute(f"SELECT * FROM specification WHERE id={project_id}").fetchall()[0]
        self.shifrkod = cursor.execute(f"SELECT shifr FROM projects WHERE id={data[1]}").fetchall()[0]
        connection.close()

        print(data)

        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        self.mtr_label = QLabel("Наименование МТР, Техническое обозначение:")
        self.mtr_edit = QLineEdit()
        self.mtr_edit.setText(data[2])
        self.layout.addWidget(self.mtr_label)
        self.layout.addWidget(self.mtr_edit)

        self.num_label = QLabel("Заводской номер или номер партии, плавки:")
        self.num_edit = QLineEdit()
        self.num_edit.setText(data[3])
        self.layout.addWidget(self.num_label)
        self.layout.addWidget(self.num_edit)

        self.factory_label = QLabel("Завод-изготовитель:")
        self.factory_edit = QLineEdit()
        self.factory_edit.setText(data[3])
        self.layout.addWidget(self.factory_label)
        self.layout.addWidget(self.factory_edit)

        self.provider_label = QLabel("Поставщик:")
        self.provider_edit = QLineEdit()
        self.provider_edit.setText(data[4])
        self.layout.addWidget(self.provider_label)
        self.layout.addWidget(self.provider_edit)

        self.count_label = QLabel("Количество:")
        self.count_edit = QLineEdit()
        self.count_edit.setText(data[5])
        self.layout.addWidget(self.count_label)
        self.layout.addWidget(self.count_edit)

        self.date_avk_label = QLabel("Дата АВК:")
        self.date_avk_edit = QLineEdit()
        self.date_avk_edit.setText(data[6])
        self.layout.addWidget(self.date_avk_label)
        self.layout.addWidget(self.date_avk_edit)

        self.status_label = QLabel("Статус:")
        self.status_edit = QLineEdit()
        self.status_edit.setText(data[7])
        self.layout.addWidget(self.status_label)
        self.layout.addWidget(self.status_edit)

        self.date_manuf_label = QLabel("Дата изготовления:")
        self.date_manuf_edit = QLineEdit()
        self.date_manuf_edit.setText(data[8])
        self.layout.addWidget(self.date_manuf_label)
        self.layout.addWidget(self.date_manuf_edit)

        self.date_delivery_label = QLabel("Дата прихода:")
        self.date_delivery_edit = QLineEdit()
        self.date_delivery_edit.setText(data[9])
        self.layout.addWidget(self.date_delivery_label)
        self.layout.addWidget(self.date_delivery_edit)

        self.gost_label = QLabel("ГОСТ, ТУ:")
        self.gost_edit = QLineEdit()
        self.gost_edit.setText(data[10])
        self.layout.addWidget(self.gost_label)
        self.layout.addWidget(self.gost_edit)

        self.transport_label = QLabel("транспортная накладная:")
        self.transport_edit = QLineEdit()
        self.transport_edit.setText(data[11])
        self.layout.addWidget(self.transport_label)
        self.layout.addWidget(self.transport_edit)

        self.passport_label = QLabel("Паспорт, сертификат:")
        self.passport_edit = QLineEdit()
        self.passport_edit.setText(data[12])
        self.layout.addWidget(self.passport_label)
        self.layout.addWidget(self.passport_edit)

        # Кнопка "Печатать документ"
        self.print_button = QPushButton("Печатать документ")
        self.print_button.clicked.connect(self.print_document)
        self.layout.addWidget(self.print_button)

    def print_document(self):
        # Здесь будет код для печати документа
        dic = {
            'имямтр': self.mtr_edit.text(),
            'заводномер': self.num_edit.text(),
            'заводизгот': self.factory_edit.text(),
            'датаизготов': self.date_avk_edit.text(),
            'постщик': self.provider_edit.text(),
            'колво': self.count_edit.text(),
            'приходдата': self.date_delivery_edit.text(),
            'госткод': self.gost_edit.text(),
            'шифркод': self.shifrkod,
            'транспорнаклад': self.transport_edit.text(),
            'пспрт': self.passport_edit.text()
        }

        doc = docx.Document(DOC_path)
        for i in doc.tables[0].rows:
            for j in i.cells:
                if j.text in dic:
                    j.text = dic[j.text]
        temp_file_path = "temp_files/temp.docx"
        doc.save(temp_file_path)

        word = win32com.client.Dispatch("Word.Application")
        doc = word.Documents.Open(os.path.dirname(__file__) + '\\temp_files\\temp.docx')

        # Сохранение документа в формате PDF
        doc.SaveAs(os.path.dirname(__file__) + '\\temp_files\\res.pdf', FileFormat=17)

        # Закрытие Word
        doc.Close()
        word.Quit()

        printer = QPrinter(QPrinter.)
        print_dialog = QPrintDialog(printer, self)
        if print_dialog.exec() == QPrintDialog.DialogCode.Accepted:
            pixmap = QPixmap('temp_files/res.pdf')
            scaled_pixmap = pixmap.scaled(QSize(printer.width(), printer.height()))
            painter = QPainter(printer)
            painter.drawPixmap(0, 0, scaled_pixmap)
            painter.end()

        os.remove(temp_file_path)
        os.remove('temp_files\\res.pdf')


class CustomInputDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Создание проекта")

        layout = QVBoxLayout(self)

        self.name_label = QLabel("Название проекта:", self)
        self.name_edit = QLineEdit(self)
        layout.addWidget(self.name_label)
        layout.addWidget(self.name_edit)

        self.shifr_label = QLabel("Шифр:", self)
        self.shifr_edit = QLineEdit(self)
        layout.addWidget(self.shifr_label)
        layout.addWidget(self.shifr_edit)

        ok_button = QPushButton("OK", self)
        ok_button.clicked.connect(self.accept)
        layout.addWidget(ok_button)

        cancel_button = QPushButton("Отмена", self)
        cancel_button.clicked.connect(self.reject)
        layout.addWidget(cancel_button)

    def get_data(self):
        return self.name_edit.text(), self.shifr_edit.text()


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Приложение для инженеров ПТО")
        self.setGeometry(100, 100, 800, 600)
        # self.setStyleSheet(''.join(open('py_dracula_dark.qss').readlines()))
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)

        button_layout = QHBoxLayout()

        add_button = QPushButton("Добавить")
        add_button.clicked.connect(self.add_project)
        button_layout.addWidget(add_button)

        delete_button = QPushButton("Удалить")
        delete_button.clicked.connect(self.delete_item)
        button_layout.addWidget(delete_button)

        self.layout = QVBoxLayout()
        self.central_widget.setLayout(self.layout)
        self.layout.addLayout(button_layout)
        self.projects_table = QTableWidget()
        self.projects_table.setColumnCount(3)
        self.projects_table.setHorizontalHeaderLabels(["ID", "Название проекта", "Шифр"])
        self.projects_table.doubleClicked.connect(self.open_project_card)

        self.layout.addWidget(self.projects_table)
        self.projects_table.resizeColumnsToContents()
        self.load_projects()

        self.projects_table.itemChanged.connect(self.update_data)

    def delete_item(self):
        selected_row = self.projects_table.currentRow()
        if selected_row >= 0:

            id_ = self.projects_table.item(selected_row, 0).text()
            connection = sqlite3.connect("projects.db")
            cursor = connection.cursor()
            cursor.execute("DELETE FROM projects WHERE id=?", (id_,))
            cursor.execute("DELETE FROM specification WHERE IDFK_project=?", (id_,))
            connection.commit()
            connection.close()
            self.load_projects()

    def load_projects(self):
        self.projects_table.setRowCount(0)
        connection = sqlite3.connect("projects.db")
        cursor = connection.cursor()
        cursor.execute("CREATE TABLE IF NOT EXISTS projects (id INTEGER PRIMARY KEY, name TEXT, shifr TEXT)")
        cursor.execute("SELECT * FROM projects")
        for row_index, row_data in enumerate(cursor.fetchall()):
            self.projects_table.insertRow(row_index)
            for col_index, col_data in enumerate(row_data):
                self.projects_table.setItem(row_index, col_index, QTableWidgetItem(str(col_data)))
        connection.commit()
        connection.close()
        self.projects_table.resizeColumnsToContents()

    def add_project(self):
        dialog = CustomInputDialog(self)
        if dialog.exec():
            name, shifr = dialog.get_data()
            connection = sqlite3.connect("projects.db")
            cursor = connection.cursor()
            cursor.execute("INSERT INTO projects (name, shifr) VALUES (?, ?)", (name, shifr))
            connection.commit()
            connection.close()
            self.load_projects()

    def open_project_card(self, item):
        project_id = self.projects_table.item(item.row(), 0).text()
        project_name = self.projects_table.item(item.row(), 1).text()
        self.project_card = ProjectCard(project_id, project_name)
        self.project_card.show()
        self.close()

    def update_data(self, item):
        connection = sqlite3.connect("projects.db")
        cursor = connection.cursor()
        row = item.row()
        column = item.column()
        new_value = item.text()
        item_id = self.projects_table.item(row, 0).text()
        column_name = self.projects_table.horizontalHeaderItem(column).text()
        teg_name = {
            "ID": 'id',
            "Название проекта": 'name',
            "Шифр": 'shifr'
        }
        cursor.execute(f"UPDATE projects SET {teg_name[column_name]}='{new_value}' WHERE id={item_id}")
        connection.commit()
        connection.close()


class ProjectCard(QDialog):
    def __init__(self, project_id, project_name):
        super().__init__()
        connection = sqlite3.connect("projects.db")
        cursor = connection.cursor()
        cursor.execute("CREATE TABLE IF NOT EXISTS specification (id INTEGER PRIMARY KEY, IDFK_project TEXT, "
                       "mtp TEXT, numb TEXT, factory TEXT, provider TEXT, count TEXT, date_AVK TEXT, status TEXT, "
                       "date_manuf TEXT, date_delivery TEXT, gost TEXT, transport_pad TEXT, "
                       "pasport TEXT)")
        data = cursor.execute(f"SELECT * FROM specification WHERE IDFK_project = {project_id}").fetchall()
        if data:
            self.setGeometry(100, 100, 800, 600)
            # self.setStyleSheet(''.join(open('py_dracula_dark.qss').readlines()))

            button_layout = QHBoxLayout()

            add_button = QPushButton("Назад")
            add_button.clicked.connect(self.go_back)
            button_layout.addWidget(add_button)

            self.layout = QVBoxLayout()
            self.setLayout(self.layout)

            self.projects_table = QTableWidget()
            self.projects_table.setColumnCount(13)
            self.projects_table.setHorizontalHeaderLabels(["ID", "Наименование МТР, Техническое обозначение",
                                                           "Заводской номер или номер партии, плавки",
                                                           "Завод-изготовитель", "Поставщик:", "Количество:",
                                                           "Дата АВК", "Статус", "Дата изготовления", "Дата прихода:",
                                                           "ГОСТ, ТУ", "транспортная накладная",
                                                           "Паспорт, сертификат"])
            self.layout.addWidget(self.projects_table)
            self.layout.addLayout(button_layout)

            for row_index, row_data in enumerate(data):
                self.projects_table.insertRow(row_index)
                for col_index, col_data in enumerate(row_data):
                    self.projects_table.setItem(row_index, col_index, QTableWidgetItem(str(col_data)))
            connection.close()
            self.projects_table.resizeColumnsToContents()
            self.load_projects()

            self.projects_table.itemChanged.connect(self.update_data)
            self.projects_table.doubleClicked.connect(self.open_project_cardform)
        else:
            self.setWindowTitle("Карточка проекта")
            self.project_id = project_id
            # self.setStyleSheet(''.join(open('py_dracula_dark.qss').readlines()))

            self.layout = QVBoxLayout()
            self.setLayout(self.layout)

            self.project_name_label = QLabel(f"Проект: {project_name}")
            self.layout.addWidget(self.project_name_label)

            self.import_button = QPushButton("Импортировать таблицу")
            self.import_button.clicked.connect(self.import_table)
            self.layout.addWidget(self.import_button)

    def import_table(self):
        file_name, _ = QFileDialog.getOpenFileName(self, "Выберите файл", "")
        print(file_name)

        if file_name:
            workbook = openpyxl.load_workbook(filename=file_name)
            sheet = workbook['ВК ЭОМ']
            connection = sqlite3.connect("projects.db")
            cursor = connection.cursor()

            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=14, values_only=True):
                if row != (1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14) and \
                        row != (None, None, None, None, None, None, None, None, None, None, None, None, None, None):
                    print(row)
                    cursor.execute(f"INSERT INTO specification (IDFK_project, mtp, numb, factory, provider, count, "
                                   f"date_AVK, status, date_manuf, date_delivery, gost, transport_pad, "
                                   f"pasport)"
                                   f"VALUES ('{self.project_id}', '{row[1]}', '{row[2]}', '{row[3]}', '{row[4]}', '{row[5]}',"
                                   f" '{row[6]}', '{row[7]}', '{row[8]}', '{row[9]}', '{row[10]}', '{row[11]}', '{row[12]}')")

            connection.commit()
            connection.close()
            al = QMessageBox.information(None, "Успех", "Операция успешно выполнена.")

            if al.Ok:
                self.main = MainWindow()
                self.main.show()
                self.close()

    def update_data(self, item):
        connection = sqlite3.connect("projects.db")
        cursor = connection.cursor()
        row = item.row()
        column = item.column()
        new_value = item.text()
        item_id = self.projects_table.item(row, 0).text()  # Получаем ID записи из первого столбца
        column_name = self.projects_table.horizontalHeaderItem(column).text()
        teg_name = {
            "ID": 'id',
            "Наименование МТР, Техническое обозначение": 'mtp',
            "Заводской номер или номер партии, плавки": 'numb',
            "Завод-изготовитель": 'factory',
            "Поставщик:": 'provider',
            "Количество:": 'count',
            "Дата АВК": 'date_AVK',
            "Статус": 'status',
            "Дата изготовления": 'date_manuf',
            "Дата прихода:": 'date_delivery',
            "ГОСТ, ТУ": 'gost',
            "транспортная накладная": 'transport_pad',
            "Паспорт, сертификат": 'pasport'
        }

        cursor.execute(f"UPDATE specification SET {teg_name[column_name]}='{new_value}' WHERE id={item_id}")
        connection.commit()
        connection.close()

    def load_projects(self):
        self.projects_table.setRowCount(0)

        connection = sqlite3.connect("projects.db")
        cursor = connection.cursor()
        cursor.execute("SELECT id, mtp, numb, factory,provider, count, date_AVK, status, date_manuf, date_delivery, "
                       "gost, transport_pad, pasport FROM specification")
        for row_index, row_data in enumerate(cursor.fetchall()):
            self.projects_table.insertRow(row_index)
            for col_index, col_data in enumerate(row_data):
                self.projects_table.setItem(row_index, col_index, QTableWidgetItem(str(col_data)))

        connection.commit()
        connection.close()
        self.projects_table.resizeColumnsToContents()

    def open_project_cardform(self, item):
        project_id = self.projects_table.item(item.row(), 0).text()
        project_name = self.projects_table.item(item.row(), 1).text()
        project_card = CardForm(project_id, project_name)
        project_card.exec()

    def go_back(self):
        self.main = MainWindow()
        self.main.show()
        self.close()

def main():
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
